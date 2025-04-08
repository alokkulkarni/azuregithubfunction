import os
import logging
import requests
import pandas as pd
import json
from datetime import datetime
from dotenv import load_dotenv
from github_insights import GitHubInsights
from sonarqube_analyzer import SonarQubeAnalyzer
from nexus_iq_analyzer import NexusIQAnalyzer
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import List, Dict, Any, Optional
import time

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

class OrgRepoScanner:
    def __init__(self, token: str, org: str, sonar_url: str, sonar_token: str, 
                 nexus_url: str, nexus_username: str, nexus_password: str):
        self.token = token
        self.org = org
        self.headers = {
            'Authorization': f'token {token}',
            'Accept': 'application/vnd.github.v3+json'
        }
        self.base_url = 'https://api.github.com'
        self.insights_client = GitHubInsights(token, org)
        self.sonar_analyzer = SonarQubeAnalyzer(sonar_url, sonar_token)
        self.nexus_analyzer = NexusIQAnalyzer(nexus_url, nexus_username, nexus_password)
        self.checkpoint_file = f"{org}_scan_checkpoint.json"
        self.results_file = f"{org}_scan_results.json"

    def load_checkpoint(self) -> tuple[Optional[int], List[Dict]]:
        """Load checkpoint data if it exists."""
        try:
            if os.path.exists(self.checkpoint_file):
                with open(self.checkpoint_file, 'r') as f:
                    checkpoint = json.load(f)
                    return checkpoint.get('last_page', 0), checkpoint.get('processed_repos', [])
            return 0, []
        except Exception as e:
            logging.error(f"Error loading checkpoint: {str(e)}")
            return 0, []

    def save_checkpoint(self, last_page: int, processed_repos: List[Dict]):
        """Save checkpoint data."""
        try:
            checkpoint_data = {
                'last_page': last_page,
                'processed_repos': processed_repos,
                'timestamp': datetime.now().isoformat()
            }
            with open(self.checkpoint_file, 'w') as f:
                json.dump(checkpoint_data, f)
        except Exception as e:
            logging.error(f"Error saving checkpoint: {str(e)}")

    def save_results(self, results: List[Dict]):
        """Save current results to file."""
        try:
            with open(self.results_file, 'w') as f:
                json.dump(results, f)
        except Exception as e:
            logging.error(f"Error saving results: {str(e)}")

    def process_repository(self, repo: Dict) -> Optional[Dict]:
        """Process a single repository with all analyzers."""
        try:
            repo_name = repo.get('name')
            if not repo_name:
                return None

            logging.info(f"Processing repository: {repo_name}")
            
            # Get GitHub insights
            insights = self.get_repo_insights(repo)
            if not insights:
                return None

            # Get SonarQube data
            project_key = repo_name.lower()
            if project_info := self.sonar_analyzer.get_project_info(project_key):
                sonar_metrics = self.sonar_analyzer.get_project_metrics(project_key)
                insights.update({
                    'SonarQube Status': 'Active',
                    'Quality Gate': sonar_metrics['quality_gate_status'],
                    'Bugs': sonar_metrics['bugs'],
                    'Vulnerabilities': sonar_metrics['vulnerabilities'],
                    'Code Smells': sonar_metrics['code_smells'],
                    'Coverage (%)': f"{sonar_metrics['coverage']:.1f}",
                    'Duplication (%)': f"{sonar_metrics['duplicated_lines_density']:.1f}",
                    'Security Rating': sonar_metrics['security_rating'],
                    'Reliability Rating': sonar_metrics['reliability_rating'],
                    'Maintainability Rating': sonar_metrics['sqale_rating'],
                    'Lines of Code': sonar_metrics['lines_of_code'],
                    'Cognitive Complexity': sonar_metrics['cognitive_complexity'],
                    'Technical Debt': sonar_metrics['technical_debt'],
                    'Test Success (%)': f"{sonar_metrics['test_success_density']:.1f}",
                    'Test Failures': sonar_metrics['test_failures'],
                    'Test Errors': sonar_metrics['test_errors'],
                    'Last Analysis': sonar_metrics['last_analysis']
                })
            else:
                insights['SonarQube Status'] = 'Not Found'

            # Get Nexus IQ data
            if app_info := self.nexus_analyzer.get_application_info(repo_name):
                nexus_metrics = self.nexus_analyzer.get_security_metrics(app_info['id'])
                insights.update({
                    'Nexus IQ Status': 'Active',
                    'Critical Issues': nexus_metrics['critical_issues'],
                    'Severe Issues': nexus_metrics['severe_issues'],
                    'Moderate Issues': nexus_metrics['moderate_issues'],
                    'Low Issues': nexus_metrics['low_issues'],
                    'Policy Violations': nexus_metrics['policy_violations'],
                    'Security Violations': nexus_metrics['security_violations'],
                    'License Violations': nexus_metrics['license_violations'],
                    'Quality Violations': nexus_metrics['quality_violations'],
                    'Total Components': nexus_metrics['total_components'],
                    'Vulnerable Components': nexus_metrics['vulnerable_components'],
                    'Risk Score': f"{nexus_metrics['risk_score']:.1f}",
                    'Policy Action': nexus_metrics['policy_action'],
                    'Last Scan': nexus_metrics['last_scan_date'],
                    'Evaluated Components': nexus_metrics['evaluated_components']
                })
            else:
                insights['Nexus IQ Status'] = 'Not Found'

            return insights

        except Exception as e:
            logging.error(f"Error processing repository {repo.get('name', 'unknown')}: {str(e)}")
            return None

    def process_page(self, repos: List[Dict]) -> List[Dict]:
        """Process a page of repositories in parallel."""
        results = []
        with ThreadPoolExecutor(max_workers=10) as executor:
            future_to_repo = {executor.submit(self.process_repository, repo): repo for repo in repos}
            for future in as_completed(future_to_repo):
                repo = future_to_repo[future]
                try:
                    result = future.result()
                    if result:
                        results.append(result)
                except Exception as e:
                    logging.error(f"Error processing {repo.get('name', 'unknown')}: {str(e)}")
        return results

    def scan_and_export(self, output_file: str) -> None:
        """Scan all repositories and export insights to Excel with parallel processing and resume capability."""
        try:
            # Load checkpoint
            last_processed_page, processed_repos = self.load_checkpoint()
            
            # Continue from where we left off
            page = last_processed_page + 1
            per_page = 100

            all_results = processed_repos
            while True:
                # Fetch repositories for current page
                url = f'{self.base_url}/orgs/{self.org}/repos'
                params = {
                    'per_page': per_page,
                    'page': page,
                    'sort': 'updated',
                    'direction': 'desc'
                }
                
                try:
                    response = requests.get(url, headers=self.headers, params=params)
                    response.raise_for_status()
                    repos = response.json()
                    
                    if not repos:
                        break
                    
                    logging.info(f"Processing page {page} with {len(repos)} repositories")
                    
                    # Process current page in parallel
                    page_results = self.process_page(repos)
                    all_results.extend(page_results)
                    
                    # Save checkpoint after each page
                    self.save_checkpoint(page, all_results)
                    self.save_results(all_results)
                    
                    if len(repos) < per_page:
                        break
                        
                    page += 1
                    
                except Exception as e:
                    logging.error(f"Error processing page {page}: {str(e)}")
                    # Save checkpoint before exiting
                    self.save_checkpoint(page - 1, all_results)
                    self.save_results(all_results)
                    raise

            # Convert results to DataFrame
            df = pd.DataFrame(all_results)

            # Save to Excel with formatting
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Repository Insights')
                
                # Get the workbook and worksheet for formatting
                wb = writer.book
                ws = wb['Repository Insights']
                
                # Format headers
                for col_num, column in enumerate(df.columns, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.column_dimensions[get_column_letter(col_num)].width = 15
                
                # Format data cells
                for row in range(2, len(df) + 2):
                    for col in range(1, len(df.columns) + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            logging.info(f"Successfully exported insights to {output_file}")
            
            # Clean up checkpoint files after successful completion
            if os.path.exists(self.checkpoint_file):
                os.remove(self.checkpoint_file)
            if os.path.exists(self.results_file):
                os.remove(self.results_file)
            
        except Exception as e:
            logging.error(f"Error in scan_and_export: {str(e)}")
            raise

def main():
    """Main function to run the repository scanner."""
    # Load environment variables
    load_dotenv()
    
    # Get required environment variables
    github_token = os.getenv('GITHUB_TOKEN')
    github_org = os.getenv('GITHUB_ORG')
    sonar_url = os.getenv('SONAR_URL', 'http://localhost:9000')
    sonar_token = os.getenv('SONAR_TOKEN')
    nexus_url = os.getenv('NEXUS_URL')
    nexus_username = os.getenv('NEXUS_USERNAME')
    nexus_password = os.getenv('NEXUS_PASSWORD')
    
    if not all([github_token, github_org, sonar_token, nexus_url, nexus_username, nexus_password]):
        logging.error("Missing required environment variables")
        return
    
    # Initialize scanner
    scanner = OrgRepoScanner(github_token, github_org, sonar_url, sonar_token,
                            nexus_url, nexus_username, nexus_password)
    
    # Create output filename with organization name
    output_file = f"{github_org}_repository_insights.xlsx"
    
    # Scan repositories and export insights
    scanner.scan_and_export(output_file)

if __name__ == "__main__":
    main() 