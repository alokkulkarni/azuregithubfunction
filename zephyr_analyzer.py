import os
import logging
import requests
import pandas as pd
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

class ZephyrAnalyzer:
    """Class to analyze test cases and executions from Zephyr Scale."""
    
    def __init__(self, jira_url: str, zephyr_token: str):
        self.base_url = jira_url.rstrip('/')
        self.headers = {
            'Authorization': f'Bearer {zephyr_token}',
            'Content-Type': 'application/json'
        }
        self.cache = {}  # Cache for API responses

    def get_project_test_cases(self, project_key: str) -> List[Dict]:
        """Get all test cases for a project."""
        try:
            if f"test_cases_{project_key}" in self.cache:
                return self.cache[f"test_cases_{project_key}"]

            test_cases = []
            start_at = 0
            max_results = 50

            while True:
                url = f"{self.base_url}/rest/atm/1.0/testcase/search"
                params = {
                    'query': f'projectKey = "{project_key}"',
                    'startAt': start_at,
                    'maxResults': max_results
                }

                response = requests.get(url, headers=self.headers, params=params)
                response.raise_for_status()
                data = response.json()

                test_cases.extend(data.get('results', []))

                if len(data.get('results', [])) < max_results:
                    break

                start_at += max_results

            self.cache[f"test_cases_{project_key}"] = test_cases
            return test_cases

        except Exception as e:
            logging.error(f"Error fetching test cases for project {project_key}: {str(e)}")
            return []

    def get_test_executions(self, project_key: str, days: int = 30) -> List[Dict]:
        """Get test executions for a project within the specified time period."""
        try:
            if f"executions_{project_key}_{days}" in self.cache:
                return self.cache[f"executions_{project_key}_{days}"]

            executions = []
            start_at = 0
            max_results = 50
            
            # Calculate date range
            end_date = datetime.now()
            start_date = end_date - timedelta(days=days)

            while True:
                url = f"{self.base_url}/rest/atm/1.0/testexecution/search"
                params = {
                    'query': (f'projectKey = "{project_key}" AND '
                             f'executedOn >= "{start_date.strftime("%Y-%m-%d")}"'),
                    'startAt': start_at,
                    'maxResults': max_results
                }

                response = requests.get(url, headers=self.headers, params=params)
                response.raise_for_status()
                data = response.json()

                executions.extend(data.get('results', []))

                if len(data.get('results', [])) < max_results:
                    break

                start_at += max_results

            self.cache[f"executions_{project_key}_{days}"] = executions
            return executions

        except Exception as e:
            logging.error(f"Error fetching test executions for project {project_key}: {str(e)}")
            return []

    def analyze_test_metrics(self, project_key: str, days: int = 30) -> Dict[str, Any]:
        """Analyze test metrics for a project."""
        metrics = {
            'total_test_cases': 0,
            'automated_tests': 0,
            'manual_tests': 0,
            'executions_last_30_days': 0,
            'passed_executions': 0,
            'failed_executions': 0,
            'blocked_executions': 0,
            'automation_coverage': 0.0,
            'test_execution_success_rate': 0.0,
            'avg_execution_time': 0.0,
            'test_cases_by_priority': {
                'High': 0,
                'Medium': 0,
                'Low': 0
            },
            'test_cases_by_type': {},
            'recent_failures': []
        }

        try:
            # Get test cases
            test_cases = self.get_project_test_cases(project_key)
            metrics['total_test_cases'] = len(test_cases)

            # Analyze test cases
            for test in test_cases:
                # Count automated vs manual tests
                if test.get('automatedTestCase', False):
                    metrics['automated_tests'] += 1
                else:
                    metrics['manual_tests'] += 1

                # Count by priority
                priority = test.get('priority', 'Medium')
                metrics['test_cases_by_priority'][priority] = metrics['test_cases_by_priority'].get(priority, 0) + 1

                # Count by type
                test_type = test.get('type', 'Functional')
                metrics['test_cases_by_type'][test_type] = metrics['test_cases_by_type'].get(test_type, 0) + 1

            # Calculate automation coverage
            if metrics['total_test_cases'] > 0:
                metrics['automation_coverage'] = (metrics['automated_tests'] / metrics['total_test_cases']) * 100

            # Get and analyze test executions
            executions = self.get_test_executions(project_key, days)
            metrics['executions_last_30_days'] = len(executions)

            total_execution_time = 0
            for execution in executions:
                status = execution.get('status', {}).get('name', '').lower()
                
                if status == 'pass':
                    metrics['passed_executions'] += 1
                elif status == 'fail':
                    metrics['failed_executions'] += 1
                    # Track recent failures
                    metrics['recent_failures'].append({
                        'test_case': execution.get('testCase', {}).get('name', 'Unknown'),
                        'execution_date': execution.get('executedOn', ''),
                        'environment': execution.get('environment', 'Unknown'),
                        'comment': execution.get('comment', '')
                    })
                elif status == 'blocked':
                    metrics['blocked_executions'] += 1

                # Calculate execution time if available
                if execution.get('executionTime'):
                    total_execution_time += float(execution['executionTime'])

            # Calculate success rate
            total_completed = metrics['passed_executions'] + metrics['failed_executions']
            if total_completed > 0:
                metrics['test_execution_success_rate'] = (metrics['passed_executions'] / total_completed) * 100

            # Calculate average execution time
            if metrics['executions_last_30_days'] > 0:
                metrics['avg_execution_time'] = total_execution_time / metrics['executions_last_30_days']

        except Exception as e:
            logging.error(f"Error analyzing test metrics for project {project_key}: {str(e)}")

        return metrics

    def enrich_github_analysis(self, github_analysis_file: str, output_file: str):
        """Enrich GitHub analysis with Zephyr test metrics."""
        try:
            # Read the GitHub analysis Excel file
            xls = pd.ExcelFile(github_analysis_file)
            sheets = {sheet_name: pd.read_excel(xls, sheet_name) for sheet_name in xls.sheet_names}
            
            # Add Zephyr metrics to the Summary sheet
            summary_df = sheets['Summary']
            
            # Add new columns for Zephyr metrics
            zephyr_columns = [
                'Test Cases',
                'Automated Tests',
                'Manual Tests',
                'Automation Coverage (%)',
                'Recent Executions',
                'Pass Rate (%)',
                'High Priority Tests',
                'Failed Tests',
                'Avg Execution Time (min)',
                'Last Execution'
            ]
            
            for col in zephyr_columns:
                summary_df[col] = 'N/A'
            
            # Update each repository with Zephyr data
            for idx, row in summary_df.iterrows():
                repo = row['Repository']
                project_key = f"{os.getenv('JIRA_PROJECT_PREFIX', 'PROJ')}_{repo.upper()}"
                
                metrics = self.analyze_test_metrics(project_key)
                
                if metrics['total_test_cases'] > 0:
                    summary_df.at[idx, 'Test Cases'] = metrics['total_test_cases']
                    summary_df.at[idx, 'Automated Tests'] = metrics['automated_tests']
                    summary_df.at[idx, 'Manual Tests'] = metrics['manual_tests']
                    summary_df.at[idx, 'Automation Coverage (%)'] = f"{metrics['automation_coverage']:.1f}"
                    summary_df.at[idx, 'Recent Executions'] = metrics['executions_last_30_days']
                    summary_df.at[idx, 'Pass Rate (%)'] = f"{metrics['test_execution_success_rate']:.1f}"
                    summary_df.at[idx, 'High Priority Tests'] = metrics['test_cases_by_priority']['High']
                    summary_df.at[idx, 'Failed Tests'] = metrics['failed_executions']
                    summary_df.at[idx, 'Avg Execution Time (min)'] = f"{metrics['avg_execution_time']:.1f}"
                    
                    # Get the most recent execution date
                    if metrics['recent_failures']:
                        last_execution = max(failure['execution_date'] for failure in metrics['recent_failures'])
                        summary_df.at[idx, 'Last Execution'] = last_execution
            
            # Create Test Analysis sheet
            test_analysis_data = []
            for idx, row in summary_df.iterrows():
                repo = row['Repository']
                project_key = f"{os.getenv('JIRA_PROJECT_PREFIX', 'PROJ')}_{repo.upper()}"
                metrics = self.analyze_test_metrics(project_key)
                
                if metrics['total_test_cases'] > 0:
                    # Add test type breakdown
                    for test_type, count in metrics['test_cases_by_type'].items():
                        test_analysis_data.append({
                            'Repository': repo,
                            'Test Type': test_type,
                            'Count': count,
                            'Automated': metrics['automated_tests'],
                            'Manual': metrics['manual_tests'],
                            'Success Rate': f"{metrics['test_execution_success_rate']:.1f}%",
                            'Recent Failures': len(metrics['recent_failures'])
                        })
            
            test_analysis_df = pd.DataFrame(test_analysis_data)
            
            # Write to Excel
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # Write updated summary sheet
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Write test analysis sheet
                if not test_analysis_df.empty:
                    test_analysis_df.to_excel(writer, sheet_name='Test Analysis', index=False)
                
                # Copy other sheets as is
                for sheet_name, df in sheets.items():
                    if sheet_name not in ['Summary', 'Test Analysis']:
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Format the sheets
                self._format_excel_sheets(writer, summary_df, test_analysis_df)
            
            logging.info(f"Successfully enriched analysis with Zephyr data: {output_file}")
            
        except Exception as e:
            logging.error(f"Error enriching GitHub analysis with Zephyr data: {str(e)}")
            raise

    def _format_excel_sheets(self, writer: pd.ExcelWriter, summary_df: pd.DataFrame, test_analysis_df: pd.DataFrame):
        """Format Excel sheets with proper styling."""
        # Format Summary sheet
        worksheet = writer.sheets['Summary']
        
        # Set column widths
        column_widths = {
            'A': 30,  # Repository
            'B': 15,  # Quality Score
            'C': 15,  # Test Cases
            'D': 15,  # Automated Tests
            'E': 15,  # Manual Tests
            'F': 20,  # Automation Coverage
            'G': 15,  # Recent Executions
            'H': 15,  # Pass Rate
            'I': 15,  # High Priority Tests
            'J': 15,  # Failed Tests
            'K': 20,  # Avg Execution Time
            'L': 20   # Last Execution
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Format headers
        header_style = {
            'font': Font(bold=True),
            'fill': PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
        }
        
        for cell in worksheet[1]:
            cell.font = header_style['font']
            cell.fill = header_style['fill']
            cell.alignment = header_style['alignment']
        
        # Format data cells
        data_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = data_alignment
        
        # Format Test Analysis sheet if it exists
        if not test_analysis_df.empty and 'Test Analysis' in writer.sheets:
            worksheet = writer.sheets['Test Analysis']
            
            # Set column widths
            for idx, col in enumerate(test_analysis_df.columns):
                worksheet.column_dimensions[chr(65 + idx)].width = 20
            
            # Format headers
            for cell in worksheet[1]:
                cell.font = header_style['font']
                cell.fill = header_style['fill']
                cell.alignment = header_style['alignment']
            
            # Format data cells
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = data_alignment

def main():
    """Main function to run the Zephyr analysis."""
    # Load environment variables
    load_dotenv()
    
    # Get required environment variables
    jira_url = os.getenv('JIRA_URL')
    zephyr_token = os.getenv('ZEPHYR_TOKEN')
    
    if not all([jira_url, zephyr_token]):
        logging.error("Missing required environment variables: JIRA_URL, ZEPHYR_TOKEN")
        return
    
    # Find the most recent GitHub analysis file
    analysis_files = [f for f in os.listdir('.') if f.endswith('.xlsx') and 'code_quality_analysis' in f]
    if not analysis_files:
        logging.error("No GitHub analysis file found")
        return
    
    # Get the most recent file
    latest_file = max(analysis_files, key=os.path.getctime)
    
    # Create output filename
    output_file = latest_file.replace('.xlsx', '_with_tests.xlsx')
    
    # Initialize Zephyr analyzer
    analyzer = ZephyrAnalyzer(jira_url, zephyr_token)
    
    # Enrich GitHub analysis with Zephyr data
    analyzer.enrich_github_analysis(latest_file, output_file)

if __name__ == "__main__":
    main() 