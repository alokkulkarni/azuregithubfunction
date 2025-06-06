import os
import logging
import requests
import pandas as pd
from datetime import datetime
from typing import Dict, List, Any, Optional
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

class SonarQubeAnalyzer:
    """Class to analyze SonarQube data for repositories."""
    
    def __init__(self, sonar_url: str, sonar_token: str):
        self.base_url = sonar_url.rstrip('/')
        self.headers = {
            'Content-Type': 'application/json',
            'Authorization': f'Bearer {sonar_token}'
        }

    def to_camel_case(self, project_key: str) -> str:
        """Convert project key to camel case by splitting on underscores and capitalizing each word."""
        words = project_key.split('-')
        return ''.join(word.capitalize() for word in words)

    def get_project_info(self, project_key: str) -> Optional[Dict]:
        """Get project information from SonarQube using measures endpoint."""
        camel_case_key = self.to_camel_case(project_key)
        try:
            url = f"{self.base_url}/api/measures/component"
            params = {
                'component': camel_case_key,
                'metricKeys': 'ncloc'  # Using a simple metric to validate project existence
            }
            response = requests.get(url, headers=self.headers, params=params)
            response.raise_for_status()
            
            data = response.json()
            component = data.get('component')
            
            # If we get a component back, the project exists
            return component if component else None
            
        except requests.exceptions.HTTPError as e:
            if e.response.status_code == 404:
                logging.info(f"Project {camel_case_key} not found in SonarQube")
            else:
                logging.error(f"Error fetching project info for {camel_case_key}: {str(e)}")
            return None
        except Exception as e:
            logging.error(f"Error fetching project info for {camel_case_key}: {str(e)}")
            return None

    def get_project_metrics(self, project_key: str) -> Dict[str, Any]:
        """Get quality metrics for a project."""
        metrics = {
            'bugs': 0,
            'vulnerabilities': 0,
            'code_smells': 0,
            'coverage': 0.0,
            'duplicated_lines_density': 0.0,
            'security_rating': 'N/A',
            'reliability_rating': 'N/A',
            'sqale_rating': 'N/A',
            'last_analysis': 'Never',
            'quality_gate_status': 'N/A',
            'lines_of_code': 0,
            'cognitive_complexity': 0,
            'technical_debt': '0min',
            'test_success_density': 0.0,
            'test_failures': 0,
            'test_errors': 0
        }

        camel_case_key = self.to_camel_case(project_key)
        
        try:
            # Get measures
            url = f"{self.base_url}/api/measures/component"
            params = {
                'component': camel_case_key,
                'metricKeys': ('bugs,vulnerabilities,code_smells,coverage,duplicated_lines_density,'
                             'security_rating,reliability_rating,sqale_rating,ncloc,cognitive_complexity,'
                             'sqale_index,test_success_density,test_failures,test_errors')
            }
            response = requests.get(url, headers=self.headers, params=params)
            response.raise_for_status()
            
            data = response.json()
            measures = {m['metric']: m['value'] for m in data.get('component', {}).get('measures', [])}
            
            # Update metrics with actual values
            metrics.update({
                'bugs': int(measures.get('bugs', 0)),
                'vulnerabilities': int(measures.get('vulnerabilities', 0)),
                'code_smells': int(measures.get('code_smells', 0)),
                'coverage': float(measures.get('coverage', 0)),
                'duplicated_lines_density': float(measures.get('duplicated_lines_density', 0)),
                'security_rating': self._convert_rating(measures.get('security_rating', '0')),
                'reliability_rating': self._convert_rating(measures.get('reliability_rating', '0')),
                'sqale_rating': self._convert_rating(measures.get('sqale_rating', '0')),
                'lines_of_code': int(measures.get('ncloc', 0)),
                'cognitive_complexity': int(measures.get('cognitive_complexity', 0)),
                'technical_debt': self._convert_technical_debt(measures.get('sqale_index', '0')),
                'test_success_density': float(measures.get('test_success_density', 0)),
                'test_failures': int(measures.get('test_failures', 0)),
                'test_errors': int(measures.get('test_errors', 0))
            })
            
            # Get quality gate status
            url = f"{self.base_url}/api/qualitygates/project_status"
            params = {'projectKey': camel_case_key}
            response = requests.get(url, headers=self.headers, params=params)
            response.raise_for_status()
            
            data = response.json()
            metrics['quality_gate_status'] = data.get('projectStatus', {}).get('status', 'N/A')
            
            # Get last analysis date
            url = f"{self.base_url}/api/project_analyses/search"
            params = {
                'project': camel_case_key,
                'ps': 1  # Get only the latest analysis
            }
            response = requests.get(url, headers=self.headers, params=params)
            response.raise_for_status()
            
            data = response.json()
            analyses = data.get('analyses', [])
            if analyses:
                metrics['last_analysis'] = analyses[0].get('date', 'N/A')
            
        except Exception as e:
            logging.error(f"Error fetching metrics for {camel_case_key}: {str(e)}")
        
        return metrics

    def _convert_rating(self, rating: str) -> str:
        """Convert SonarQube rating from number to letter grade."""
        rating_map = {
            '1': 'A',
            '2': 'B',
            '3': 'C',
            '4': 'D',
            '5': 'E'
        }
        return rating_map.get(rating, 'N/A')

    def _convert_technical_debt(self, minutes: str) -> str:
        """Convert technical debt from minutes to readable format."""
        try:
            minutes = int(minutes)
            if minutes < 60:
                return f"{minutes}min"
            elif minutes < 1440:  # 24 hours
                hours = minutes // 60
                return f"{hours}h {minutes % 60}min"
            else:
                days = minutes // 1440
                hours = (minutes % 1440) // 60
                return f"{days}d {hours}h"
        except (ValueError, TypeError):
            return "0min"

    def update_excel_with_sonarqube_data(self, excel_file: str):
        """Update Excel file with SonarQube analysis data."""
        try:
            # Read the Excel file directly
            df = pd.read_excel(excel_file, engine='openpyxl')
            
            if 'Repository' not in df.columns:
                raise ValueError("Could not find 'Repository' column in the Excel file")
            
            # Define SonarQube columns to add
            sonar_columns = [
                'SonarQube Status',
                'Quality Gate',
                'Bugs',
                'Vulnerabilities',
                'Code Smells',
                'Coverage (%)',
                'Duplication (%)',
                'Security Rating',
                'Reliability Rating',
                'Maintainability Rating',
                'Lines of Code',
                'Cognitive Complexity',
                'Technical Debt',
                'Test Success (%)',
                'Test Failures',
                'Test Errors',
                'Last Analysis'
            ]
            
            # Initialize new columns with 'N/A'
            for col in sonar_columns:
                df[col] = 'N/A'
            
            # Process each repository
            for idx, row in df.iterrows():
                repo = row['Repository']
                if pd.notna(repo):  # Check if repository name is not NaN
                    project_key = f"{repo}".lower()
                    camel_case_key = self.to_camel_case(project_key)
                    logging.info(f"Processing repository: {repo} (Project key: {camel_case_key})")
                    
                    if project_info := self.get_project_info(project_key):
                        metrics = self.get_project_metrics(project_key)
                        
                        # Update DataFrame with metrics
                        df.at[idx, 'SonarQube Status'] = 'Active'
                        df.at[idx, 'Quality Gate'] = metrics['quality_gate_status']
                        df.at[idx, 'Bugs'] = metrics['bugs']
                        df.at[idx, 'Vulnerabilities'] = metrics['vulnerabilities']
                        df.at[idx, 'Code Smells'] = metrics['code_smells']
                        df.at[idx, 'Coverage (%)'] = f"{metrics['coverage']:.1f}"
                        df.at[idx, 'Duplication (%)'] = f"{metrics['duplicated_lines_density']:.1f}"
                        df.at[idx, 'Security Rating'] = metrics['security_rating']
                        df.at[idx, 'Reliability Rating'] = metrics['reliability_rating']
                        df.at[idx, 'Maintainability Rating'] = metrics['sqale_rating']
                        df.at[idx, 'Lines of Code'] = metrics['lines_of_code']
                        df.at[idx, 'Cognitive Complexity'] = metrics['cognitive_complexity']
                        df.at[idx, 'Technical Debt'] = metrics['technical_debt']
                        df.at[idx, 'Test Success (%)'] = f"{metrics['test_success_density']:.1f}"
                        df.at[idx, 'Test Failures'] = metrics['test_failures']
                        df.at[idx, 'Test Errors'] = metrics['test_errors']
                        df.at[idx, 'Last Analysis'] = metrics['last_analysis']
                    else:
                        df.at[idx, 'SonarQube Status'] = 'Not Found'
            
            # Save directly to a new file first
            output_file = f"{os.path.splitext(excel_file)[0]}_new.xlsx"
            
            # Save with pandas first
            df.to_excel(output_file, index=False, engine='openpyxl')
            
            # Now apply formatting
            wb = openpyxl.load_workbook(output_file)
            ws = wb.active
            
            # Format headers
            for col_num, header in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.column_dimensions[get_column_letter(col_num)].width = 15
            
            # Format data cells
            for row in range(2, len(df) + 2):
                for col in range(1, len(df.columns) + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Save the formatted workbook
            wb.save(output_file)
            wb.close()
            
            # If everything was successful, replace the original file
            if os.path.exists(excel_file):
                os.remove(excel_file)
            os.rename(output_file, excel_file)
            logging.info(f"Successfully updated {excel_file} with SonarQube data")
            
        except Exception as e:
            logging.error(f"Error updating Excel file: {str(e)}")
            # Clean up output file if it exists
            if 'output_file' in locals() and os.path.exists(output_file):
                try:
                    os.remove(output_file)
                except Exception as cleanup_error:
                    logging.error(f"Error cleaning up output file: {cleanup_error}")
            raise

def main():
    """Main function to run the SonarQube analysis."""
    # Load environment variables
    load_dotenv()
    
    # Get required environment variables
    sonar_url = os.getenv('SONAR_URL', 'http://localhost:9000')
    sonar_token = os.getenv('SONAR_TOKEN')
    github_org = os.getenv('GITHUB_ORG')
    
    if not sonar_token:
        logging.error("Missing required environment variable: SONAR_TOKEN")
        return
    
    if not github_org:
        logging.error("Missing required environment variable: GITHUB_ORG")
        return
    
    # Look for the Excel file with org name pattern
    excel_filename = f"{github_org}_repository_insights.xlsx"
    
    if not os.path.exists(excel_filename):
        logging.error(f"Excel file not found: {excel_filename}")
        return
    
    logging.info(f"Found analysis file: {excel_filename}")
    
    # Initialize SonarQube analyzer
    analyzer = SonarQubeAnalyzer(sonar_url, sonar_token)
    
    # Update the Excel file with SonarQube data
    analyzer.update_excel_with_sonarqube_data(excel_filename)

if __name__ == "__main__":
    main() 