import os
import logging
import requests
import pandas as pd
from datetime import datetime, UTC, timedelta
from dotenv import load_dotenv
from typing import Dict, List, Any
import math
import openpyxl

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

class IndustryStandards:
    """Industry standards and benchmarks for code quality metrics."""
    
    COMMIT_STANDARDS = {
        'frequency': {
            'excellent': {'min_weekly': 3, 'max_weekly': 15, 'variance_threshold': 5},
            'good': {'min_weekly': 2, 'max_weekly': 20, 'variance_threshold': 10},
            'average': {'min_weekly': 1, 'max_weekly': 25, 'variance_threshold': 15},
            'below_average': {'min_weekly': 0.5, 'max_weekly': 30, 'variance_threshold': 20}
        },
        'description': {
            'excellent': 'Consistent, regular commits indicating continuous integration practices',
            'good': 'Regular commits with acceptable variance',
            'average': 'Moderate commit frequency with some inconsistency',
            'below_average': 'Irregular commit patterns indicating potential process issues'
        }
    }
    
    CODE_CHURN_STANDARDS = {
        'weekly_churn': {
            'excellent': {'ratio': 200, 'deletion_ratio': 0.8},
            'good': {'ratio': 500, 'deletion_ratio': 1.0},
            'average': {'ratio': 1000, 'deletion_ratio': 1.2},
            'below_average': {'ratio': 2000, 'deletion_ratio': 1.5}
        },
        'description': {
            'excellent': 'Low churn indicating stable, well-planned development',
            'good': 'Moderate churn with balanced additions/deletions',
            'average': 'Notable churn but within acceptable limits',
            'below_average': 'High churn indicating potential stability issues'
        }
    }
    
    BRANCH_STANDARDS = {
        'complexity': {
            'excellent': {'max_branches': 5, 'max_age_days': 7},
            'good': {'max_branches': 8, 'max_age_days': 14},
            'average': {'max_branches': 12, 'max_age_days': 30},
            'below_average': {'max_branches': 15, 'max_age_days': 60}
        },
        'description': {
            'excellent': 'Clean branch strategy with quick integration',
            'good': 'Well-managed branches with timely merging',
            'average': 'Acceptable branch count with some stale branches',
            'below_average': 'Too many branches or long-lived feature branches'
        }
    }
    
    CODE_QUALITY_STANDARDS = {
        'test_coverage': {
            'excellent': {'min_ratio': 0.8, 'min_coverage': 90},
            'good': {'min_ratio': 0.6, 'min_coverage': 80},
            'average': {'min_ratio': 0.4, 'min_coverage': 70},
            'below_average': {'min_ratio': 0.2, 'min_coverage': 50}
        },
        'documentation': {
            'excellent': {'doc_ratio': 0.2, 'has_wiki': True, 'has_contributing': True},
            'good': {'doc_ratio': 0.15, 'has_wiki': True, 'has_contributing': False},
            'average': {'doc_ratio': 0.1, 'has_wiki': False, 'has_contributing': False},
            'below_average': {'doc_ratio': 0.05, 'has_wiki': False, 'has_contributing': False}
        }
    }
    
    ABERRANCY_STANDARDS = {
        'excellent': {
            'score_range': (0, 20),
            'description': 'Minimal deviation from best practices',
            'risk_level': 'Low Risk'
        },
        'good': {
            'score_range': (20, 40),
            'description': 'Minor deviations from best practices',
            'risk_level': 'Moderate Risk'
        },
        'average': {
            'score_range': (40, 60),
            'description': 'Notable deviations from best practices',
            'risk_level': 'Medium Risk'
        },
        'below_average': {
            'score_range': (60, 100),
            'description': 'Significant deviations from best practices',
            'risk_level': 'High Risk'
        }
    }
    
    @staticmethod
    def get_commit_rating(avg_weekly_commits: float, variance: float) -> Dict[str, Any]:
        """Get rating based on commit patterns."""
        for rating, criteria in IndustryStandards.COMMIT_STANDARDS['frequency'].items():
            if (criteria['min_weekly'] <= avg_weekly_commits <= criteria['max_weekly'] and 
                variance <= criteria['variance_threshold']):
                return {
                    'rating': rating,
                    'description': IndustryStandards.COMMIT_STANDARDS['description'][rating],
                    'industry_avg': criteria['min_weekly'],
                    'variance_threshold': criteria['variance_threshold']
                }
        return {
            'rating': 'below_average',
            'description': IndustryStandards.COMMIT_STANDARDS['description']['below_average'],
            'industry_avg': IndustryStandards.COMMIT_STANDARDS['frequency']['average']['min_weekly'],
            'variance_threshold': IndustryStandards.COMMIT_STANDARDS['frequency']['average']['variance_threshold']
        }

    @staticmethod
    def get_churn_rating(weekly_churn: float, deletion_ratio: float) -> Dict[str, Any]:
        """Get rating based on code churn."""
        for rating, criteria in IndustryStandards.CODE_CHURN_STANDARDS['weekly_churn'].items():
            if weekly_churn <= criteria['ratio'] and deletion_ratio <= criteria['deletion_ratio']:
                return {
                    'rating': rating,
                    'description': IndustryStandards.CODE_CHURN_STANDARDS['description'][rating],
                    'industry_threshold': criteria['ratio'],
                    'deletion_ratio_threshold': criteria['deletion_ratio']
                }
        return {
            'rating': 'below_average',
            'description': IndustryStandards.CODE_CHURN_STANDARDS['description']['below_average'],
            'industry_threshold': IndustryStandards.CODE_CHURN_STANDARDS['weekly_churn']['average']['ratio'],
            'deletion_ratio_threshold': IndustryStandards.CODE_CHURN_STANDARDS['weekly_churn']['average']['deletion_ratio']
        }

    @staticmethod
    def get_branch_rating(branch_count: int, max_branch_age_days: float) -> Dict[str, Any]:
        """Get rating based on branch complexity."""
        for rating, criteria in IndustryStandards.BRANCH_STANDARDS['complexity'].items():
            if branch_count <= criteria['max_branches'] and max_branch_age_days <= criteria['max_age_days']:
                return {
                    'rating': rating,
                    'description': IndustryStandards.BRANCH_STANDARDS['description'][rating],
                    'industry_max_branches': criteria['max_branches'],
                    'industry_max_age': criteria['max_age_days']
                }
        return {
            'rating': 'below_average',
            'description': IndustryStandards.BRANCH_STANDARDS['description']['below_average'],
            'industry_max_branches': IndustryStandards.BRANCH_STANDARDS['complexity']['average']['max_branches'],
            'industry_max_age': IndustryStandards.BRANCH_STANDARDS['complexity']['average']['max_age_days']
        }

    @staticmethod
    def get_aberrancy_rating(aberrancy_score: float) -> Dict[str, Any]:
        """Get rating based on aberrancy score."""
        for rating, criteria in IndustryStandards.ABERRANCY_STANDARDS.items():
            min_score, max_score = criteria['score_range']
            if min_score <= aberrancy_score < max_score:
                return {
                    'rating': rating,
                    'description': criteria['description'],
                    'risk_level': criteria['risk_level']
                }
        return {
            'rating': 'below_average',
            'description': IndustryStandards.ABERRANCY_STANDARDS['below_average']['description'],
            'risk_level': IndustryStandards.ABERRANCY_STANDARDS['below_average']['risk_level']
        }

class MetricDefinitions:
    """Definitions and explanations for all metrics used in the analysis."""
    
    QUALITY_METRICS = {
        'quality_score': 'Overall code quality score (0-100) based on various factors including CI/CD, tests, and documentation',
        'total_files': 'Total number of files in the repository',
        'code_files': 'Number of files containing actual source code',
        'test_files': 'Number of test files in the repository',
        'doc_files': 'Number of documentation files',
        'large_files': 'Number of files larger than 1MB (potential code smell)',
        'has_ci': 'Presence of CI/CD configuration files',
        'has_tests': 'Presence of automated tests',
        'has_docs': 'Presence of documentation',
        'has_license': 'Presence of a license file',
        'code_to_test_ratio': 'Ratio of code files to test files (lower is better)'
    }
    
    ABERRANCY_METRICS = {
        'commit_frequency': {
            'title': 'Commit Frequency Analysis',
            'description': 'Measures how often and consistently code is committed',
            'metrics': {
                'weekly_average': 'Average number of commits per week',
                'variance': 'Statistical variance in commit frequency (lower is better)',
                'score': 'Score based on commit patterns (0-100)'
            }
        },
        'code_churn': {
            'title': 'Code Churn Analysis',
            'description': 'Measures the rate of code changes and stability',
            'metrics': {
                'weekly_churn': 'Average lines of code changed per week',
                'deletion_ratio': 'Ratio of deleted lines to added lines',
                'score': 'Score based on code stability (0-100)'
            }
        },
        'branch_complexity': {
            'title': 'Branch Complexity Analysis',
            'description': 'Measures repository branching strategy and maintenance',
            'metrics': {
                'branch_count': 'Number of active branches',
                'max_age': 'Age of the oldest branch in days',
                'score': 'Score based on branch management (0-100)'
            }
        }
    }
    
    EFFORT_METRICS = {
        'total_commits': 'Total number of commits across all contributors',
        'total_changes': 'Total number of lines changed (additions + deletions)',
        'contributors': 'Number of unique contributors',
        'estimated_hours': 'Estimated development hours based on changes',
        'complexity_factor': 'Project complexity multiplier',
        'billable_hours': 'Calculated billable hours (estimated_hours × complexity_factor)'
    }
    
    RATINGS = {
        'Excellent': 'Top 10% - Follows best practices with high consistency',
        'Good': 'Top 25% - Generally well-maintained with minor improvements needed',
        'Average': 'Top 50% - Acceptable but has room for improvement',
        'Below Average': 'Bottom 50% - Needs significant improvements'
    }

class CodeQualityAnalyzer:
    def __init__(self, token: str, org: str):
        self.token = token
        self.org = org
        self.headers = {
            'Authorization': f'token {token}',
            'Accept': 'application/vnd.github.v3+json'
        }
        self.base_url = 'https://api.github.com'
        self.industry_standards = IndustryStandards()
        self.metric_definitions = MetricDefinitions()

    def get_repo_contents(self, repo: str, path: str = '') -> List[Dict]:
        """Recursively get repository contents."""
        url = f'{self.base_url}/repos/{self.org}/{repo}/contents/{path}'
        try:
            response = requests.get(url, headers=self.headers)
            response.raise_for_status()
            return response.json()
        except Exception as e:
            logging.error(f"Error fetching contents for {repo}/{path}: {str(e)}")
            return []

    def get_commit_activity(self, repo: str) -> List[Dict]:
        """Get commit activity for the past year."""
        url = f'{self.base_url}/repos/{self.org}/{repo}/stats/commit_activity'
        try:
            response = requests.get(url, headers=self.headers)
            response.raise_for_status()
            return response.json()
        except Exception as e:
            logging.error(f"Error fetching commit activity for {repo}: {str(e)}")
            return []

    def calculate_code_quality_score(self, repo: str) -> Dict[str, Any]:
        """Calculate code quality score based on various metrics."""
        quality_metrics = {
            'total_files': 0,
            'code_files': 0,
            'test_files': 0,
            'doc_files': 0,
            'large_files': 0,
            'binary_files': 0,
            'quality_score': 0,
            'has_ci': False,
            'has_tests': False,
            'has_docs': False,
            'has_license': False,
            'code_to_test_ratio': 0
        }

        # Check for important files and directories
        contents = self.get_repo_contents(repo)
        
        # File patterns to check
        code_extensions = {
            # Web Development
            '.html', '.htm', '.xhtml', '.css', '.scss', '.sass', '.less', '.styl',
            '.jsx', '.tsx', '.vue', '.svelte', '.astro', '.liquid', '.pug', '.jade',
            '.haml', '.ejs', '.hbs', '.handlebars', '.twig',
            
            # JavaScript/TypeScript
            '.js', '.ts', '.mjs', '.cjs', '.jsx', '.tsx', '.coffee', '.ls',
            '.es', '.es6', '.json', '.jsonc', '.json5',
            
            # Python
            '.py', '.pyi', '.pyx', '.pxd', '.pxi', '.pyc', '.pyd', '.pyw',
            '.ipynb', '.rpy', '.pyz', '.pyzw',
            
            # Java/Kotlin/Scala/Groovy
            '.java', '.class', '.jar', '.kt', '.kts', '.ktm',
            '.scala', '.sc', '.groovy', '.gvy', '.gy', '.gsh',
            
            # C/C++
            '.c', '.cpp', '.cc', '.cxx', '.c++', '.h', '.hpp', '.hh', '.hxx',
            '.h++', '.m', '.mm', '.inc', '.inl', '.ipp',
            
            # C#/.NET
            '.cs', '.csx', '.vb', '.fs', '.fsx', '.fsi', '.fsscript',
            '.xaml', '.razor', '.cshtml', '.vbhtml', '.aspx', '.ascx',
            
            # Ruby
            '.rb', '.rbw', '.rake', '.gemspec', '.ru', '.erb', '.rhtml',
            '.rjs', '.rxml', '.builder', '.jbuilder',
            
            # PHP
            '.php', '.php3', '.php4', '.php5', '.php7', '.phtml', '.phps',
            '.phpt', '.phar', '.inc',
            
            # Go
            '.go', '.mod', '.sum', '.tmpl', '.gohtml',
            
            # Rust
            '.rs', '.rlib', '.rst',
            
            # Swift/Objective-C
            '.swift', '.m', '.mm', '.h', '.metal',
            
            # Shell/Bash
            '.sh', '.bash', '.command', '.zsh', '.fish', '.ksh', '.csh',
            '.tcsh', '.rc', '.profile', '.bats',
            
            # Dart/Flutter
            '.dart', '.freezed.dart', '.g.dart',
            
            # SQL
            '.sql', '.mysql', '.pgsql', '.tsql', '.plsql', '.db2',
            
            # Configuration/Build
            '.xml', '.yaml', '.yml', '.toml', '.ini', '.cfg', '.conf',
            '.properties', '.env', '.gradle', '.pom', '.ant',
            
            # Mobile Development
            '.swift', '.kt', '.java', '.m', '.h', '.mm', '.dart',
            '.xcodeproj', '.pbxproj', '.storyboard', '.xib',
            
            # Systems Programming
            '.asm', '.s', '.nasm', '.masm', '.gas',
            
            # Other Languages
            '.r', '.rmd',  # R
            '.pl', '.pm', '.t',  # Perl
            '.lua',  # Lua
            '.ex', '.exs',  # Elixir
            '.erl', '.hrl',  # Erlang
            '.elm',  # Elm
            '.clj', '.cljs', '.cljc', '.edn',  # Clojure
            '.hs', '.lhs',  # Haskell
            '.ml', '.mli', '.mll', '.mly',  # OCaml
            '.f', '.f90', '.f95', '.f03', '.f08',  # Fortran
            '.mat', '.fig', '.m',  # MATLAB
            '.jl',  # Julia
            '.v', '.vh', '.sv', '.svh',  # Verilog/SystemVerilog
            '.vhd', '.vhdl',  # VHDL
            '.tcl', '.tk', '.itk',  # Tcl/Tk
            '.pro', '.pri',  # Qt/QMake
            '.cmake', '.cmake.in',  # CMake
            '.nim', '.nims',  # Nim
            '.d',  # D
            '.zig',  # Zig
            '.cr',  # Crystal
            '.rs',  # Rust
            
            # Documentation
            '.md', '.markdown', '.rst', '.adoc', '.asciidoc', '.tex',
            '.wiki', '.mediawiki', '.org',
            
            # Template Files
            '.tmpl', '.template', '.j2', '.jinja', '.jinja2',
            '.mustache', '.handlebars', '.hbs', '.ejs',
            
            # Data Formats
            '.proto', '.thrift', '.avsc', '.graphql', '.gql',
            
            # Infrastructure as Code
            '.tf', '.tfvars', '.hcl',  # Terraform
            '.cf', '.cft',  # CloudFormation
            '.k8s', '.helm',  # Kubernetes
            
            # AI/ML
            '.ipynb', '.pkl', '.h5', '.onnx', '.pbtxt', '.pb'
        }
        test_patterns = {'test', 'spec', '_test', '_spec', 'tests', 'specs'}
        doc_patterns = {'docs', 'documentation', 'wiki', 'README', 'CONTRIBUTING'}
        ci_patterns = {'.github/workflows', '.travis.yml', 'azure-pipelines.yml', 'Jenkinsfile'}
        
        for item in contents:
            if item['type'] == 'file':
                quality_metrics['total_files'] += 1
                name = item['name'].lower()
                ext = os.path.splitext(name)[1]
                
                # Check file types
                if ext in code_extensions:
                    quality_metrics['code_files'] += 1
                if any(pattern in name for pattern in test_patterns):
                    quality_metrics['test_files'] += 1
                if any(pattern in name for pattern in doc_patterns):
                    quality_metrics['doc_files'] += 1
                if item['size'] > 1000000:  # Files larger than 1MB
                    quality_metrics['large_files'] += 1
                
                # Check for specific files
                if name == 'license' or name == 'license.md':
                    quality_metrics['has_license'] = True
                
            elif item['type'] == 'dir':
                if any(pattern in item['path'] for pattern in ci_patterns):
                    quality_metrics['has_ci'] = True
                if any(pattern in item['path'].lower() for pattern in test_patterns):
                    quality_metrics['has_tests'] = True
                if any(pattern in item['path'].lower() for pattern in doc_patterns):
                    quality_metrics['has_docs'] = True

        # Calculate code to test ratio
        if quality_metrics['test_files'] > 0:
            quality_metrics['code_to_test_ratio'] = quality_metrics['code_files'] / quality_metrics['test_files']

        # Calculate quality score (0-100)
        score = 0
        score += 20 if quality_metrics['has_ci'] else 0
        score += 20 if quality_metrics['has_tests'] else 0
        score += 15 if quality_metrics['has_docs'] else 0
        score += 10 if quality_metrics['has_license'] else 0
        score += min(20, (quality_metrics['test_files'] / max(1, quality_metrics['code_files'])) * 20)
        score -= min(15, quality_metrics['large_files'] * 3)  # Penalty for large files
        
        quality_metrics['quality_score'] = max(0, min(100, score))
        
        return quality_metrics

    def calculate_aberrancy_score(self, repo: str) -> Dict[str, Any]:
        """Calculate aberrancy score based on deviations from best practices."""
        aberrancy_metrics = {
            'commit_frequency_score': 0,
            'code_churn_score': 0,
            'branch_complexity_score': 0,
            'overall_aberrancy_score': 0,
            'risk_factors': [],
            'assessment_details': {
                'commit_frequency': {
                    'score': 0,
                    'details': '',
                    'recommendations': [],
                    'industry_comparison': {}
                },
                'code_churn': {
                    'score': 0,
                    'details': '',
                    'recommendations': [],
                    'industry_comparison': {}
                },
                'branch_patterns': {
                    'score': 0,
                    'details': '',
                    'recommendations': [],
                    'industry_comparison': {}
                }
            }
        }

        # Calculate commit frequency metrics
        if commit_activity := self.get_commit_activity(repo):
            weekly_commits = [week['total'] for week in commit_activity]
            avg_commits = sum(weekly_commits) / len(weekly_commits)
            commit_variance = sum((x - avg_commits) ** 2 for x in weekly_commits) / len(weekly_commits)
            
            # Get industry comparison
            commit_rating = IndustryStandards.get_commit_rating(avg_commits, commit_variance)
            
            # Update assessment with industry comparison
            assessment = aberrancy_metrics['assessment_details']['commit_frequency']
            assessment['industry_comparison'] = {
                'rating': commit_rating['rating'],
                'description': commit_rating['description'],
                'your_weekly_avg': f"{avg_commits:.1f}",
                'industry_avg': f"{commit_rating['industry_avg']:.1f}",
                'your_variance': f"{commit_variance:.1f}",
                'industry_variance_threshold': f"{commit_rating['variance_threshold']:.1f}"
            }
            
            # Calculate score based on industry standards
            commit_freq_score = min(100, (avg_commits * 10) / (1 + math.sqrt(commit_variance)))
            aberrancy_metrics['commit_frequency_score'] = commit_freq_score
            
            assessment['score'] = commit_freq_score
            assessment['details'] = (
                f"Average commits per week: {avg_commits:.1f} "
                f"(Industry: {commit_rating['industry_avg']:.1f}), "
                f"Variance: {commit_variance:.1f} "
                f"(Industry max: {commit_rating['variance_threshold']:.1f})"
            )

        # Calculate code churn metrics
        try:
            if churn_data := self.get_code_frequency_stats(repo):
                total_additions = sum(week[1] for week in churn_data)
                total_deletions = sum(abs(week[2]) for week in churn_data)
                weeks = len(churn_data)
                weekly_churn = (total_additions + total_deletions) / weeks
                deletion_ratio = total_deletions / max(1, total_additions)
                
                # Get industry comparison
                churn_rating = IndustryStandards.get_churn_rating(weekly_churn, deletion_ratio)
                
                # Update assessment with industry comparison
                assessment = aberrancy_metrics['assessment_details']['code_churn']
                assessment['industry_comparison'] = {
                    'rating': churn_rating['rating'],
                    'description': churn_rating['description'],
                    'your_weekly_churn': f"{weekly_churn:.1f}",
                    'industry_threshold': f"{churn_rating['industry_threshold']:.1f}",
                    'your_deletion_ratio': f"{deletion_ratio:.2f}",
                    'industry_deletion_ratio': f"{churn_rating['deletion_ratio_threshold']:.2f}"
                }
                
                # Calculate score based on industry standards
                churn_score = max(0, 100 - (weekly_churn / churn_rating['industry_threshold']) * 100)
                aberrancy_metrics['code_churn_score'] = churn_score
                
                assessment['score'] = churn_score
                assessment['details'] = (
                    f"Weekly churn: {weekly_churn:.1f} lines "
                    f"(Industry max: {churn_rating['industry_threshold']:.1f}), "
                    f"Deletion ratio: {deletion_ratio:.2f} "
                    f"(Industry max: {churn_rating['deletion_ratio_threshold']:.2f})"
                )
        except Exception as e:
            logging.error(f"Error calculating code churn for {repo}: {str(e)}")

        # Calculate branch complexity metrics
        try:
            branches = self.get_branches(repo)
            if branches:
                branch_count = len(branches)
                
                # Calculate branch age
                now = datetime.now(UTC)
                branch_ages = []
                for branch in branches:
                    if commit_info := self.get_branch_last_commit(repo, branch['name']):
                        last_commit_date = datetime.strptime(
                            commit_info['commit']['committer']['date'],
                            '%Y-%m-%dT%H:%M:%SZ'
                        ).replace(tzinfo=UTC)
                        age_days = (now - last_commit_date).days
                        branch_ages.append(age_days)
                
                max_branch_age = max(branch_ages) if branch_ages else 0
                
                # Get industry comparison
                branch_rating = IndustryStandards.get_branch_rating(branch_count, max_branch_age)
                
                # Update assessment with industry comparison
                assessment = aberrancy_metrics['assessment_details']['branch_patterns']
                assessment['industry_comparison'] = {
                    'rating': branch_rating['rating'],
                    'description': branch_rating['description'],
                    'your_branch_count': str(branch_count),
                    'industry_max_branches': str(branch_rating['industry_max_branches']),
                    'your_max_age': f"{max_branch_age} days",
                    'industry_max_age': f"{branch_rating['industry_max_age']} days"
                }
                
                # Calculate score based on industry standards
                branch_score = max(0, 100 - (
                    (branch_count / branch_rating['industry_max_branches'] * 50) +
                    (max_branch_age / branch_rating['industry_max_age'] * 50)
                ))
                aberrancy_metrics['branch_complexity_score'] = branch_score
                
                assessment['score'] = branch_score
                assessment['details'] = (
                    f"Active branches: {branch_count} "
                    f"(Industry max: {branch_rating['industry_max_branches']}), "
                    f"Oldest branch: {max_branch_age} days "
                    f"(Industry max: {branch_rating['industry_max_age']} days)"
                )
        except Exception as e:
            logging.error(f"Error calculating branch complexity for {repo}: {str(e)}")

        # Calculate overall aberrancy score
        aberrancy_metrics['overall_aberrancy_score'] = 100 - (
            aberrancy_metrics['commit_frequency_score'] * 0.4 +
            aberrancy_metrics['code_churn_score'] * 0.3 +
            aberrancy_metrics['branch_complexity_score'] * 0.3
        )

        # Get aberrancy rating and risk factors
        aberrancy_rating = IndustryStandards.get_aberrancy_rating(aberrancy_metrics['overall_aberrancy_score'])
        
        # Calculate risk factors based on scores and thresholds
        risk_factors = []
        
        # Commit frequency risks
        if aberrancy_metrics['commit_frequency_score'] < 40:
            risk_factors.append("Irregular commit patterns indicating potential process issues")
        
        # Code churn risks
        if aberrancy_metrics['code_churn_score'] < 40:
            risk_factors.append("High code churn suggesting potential instability")
        
        # Branch complexity risks
        if aberrancy_metrics['branch_complexity_score'] < 40:
            risk_factors.append("Complex branching strategy with potential integration challenges")
            
        # Add specific risk factors based on assessment details
        for category, details in aberrancy_metrics['assessment_details'].items():
            comp = details.get('industry_comparison', {})
            if category == 'commit_frequency':
                if comp.get('your_variance') and float(comp['your_variance'].replace('N/A', '0')) > float(comp.get('industry_variance_threshold', '10')):
                    risk_factors.append("High variance in commit frequency")
            elif category == 'code_churn':
                if comp.get('your_deletion_ratio') and float(comp['your_deletion_ratio'].replace('N/A', '0')) > float(comp.get('industry_deletion_ratio', '1.0')):
                    risk_factors.append("High code deletion ratio")
            elif category == 'branch_patterns':
                if comp.get('your_max_age') and 'days' in comp['your_max_age']:
                    days = float(comp['your_max_age'].split()[0])
                    if days > float(comp.get('industry_max_age', '30').split()[0]):
                        risk_factors.append("Long-lived branches detected")

        aberrancy_metrics['risk_factors'] = risk_factors
        aberrancy_metrics['aberrancy_rating'] = aberrancy_rating

        return aberrancy_metrics

    def calculate_billable_efforts(self, repo: str) -> Dict[str, Any]:
        """Calculate billable coding efforts based on commits and changes."""
        effort_metrics = {
            'total_commits': 0,
            'total_changes': 0,
            'contributors': 0,
            'estimated_hours': 0,
            'complexity_factor': 1.0,
            'billable_hours': 0
        }

        try:
            url = f'{self.base_url}/repos/{self.org}/{repo}/stats/contributors'
            response = requests.get(url, headers=self.headers)
            response.raise_for_status()
            stats = response.json()

            if not isinstance(stats, list):
                logging.error(f"Invalid contributor stats format for {repo}")
                return effort_metrics

            for contributor in stats:
                if not isinstance(contributor, dict):
                    continue
                    
                effort_metrics['contributors'] += 1
                effort_metrics['total_commits'] += contributor.get('total', 0)
                
                # Calculate total changes
                weeks = contributor.get('weeks', [])
                if isinstance(weeks, list):
                    for week in weeks:
                        if isinstance(week, dict):
                            effort_metrics['total_changes'] += (
                                week.get('additions', 0) + week.get('deletions', 0)
                            )

            # Estimate hours based on changes and complexity
            effort_metrics['estimated_hours'] = effort_metrics['total_changes'] / 100  # Assume 100 changes per hour
            effort_metrics['billable_hours'] = effort_metrics['estimated_hours'] * effort_metrics['complexity_factor']

        except Exception as e:
            logging.error(f"Error calculating billable efforts for {repo}: {str(e)}")

        return effort_metrics

    def analyze_repository(self, repo: str) -> Dict[str, Any]:
        """Analyze a repository for all metrics."""
        logging.info(f"Analyzing repository: {repo}")
        
        analysis = {
            'repository': repo,
            'analyzed_at': datetime.now(UTC).isoformat(),
            'quality_metrics': self.calculate_code_quality_score(repo),
            'aberrancy_metrics': self.calculate_aberrancy_score(repo),
            'effort_metrics': self.calculate_billable_efforts(repo)
        }
        
        return analysis

    def export_to_excel(self, analyses: List[Dict], output_file: str):
        """Export analyses to Excel with multiple sheets."""
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            self._export_summary(analyses, writer)
            self._export_industry_standards(writer)
            self._export_metric_definitions(writer)
            self._export_detailed_metrics(analyses, writer)
            self._format_excel_sheets(writer)

    def _export_summary(self, analyses: List[Dict], writer: pd.ExcelWriter):
        """Export summary data to Excel."""
        summary_data = []
        for analysis in analyses:
            aberrancy_score = analysis['aberrancy_metrics']['overall_aberrancy_score']
            aberrancy_rating = analysis['aberrancy_metrics'].get('aberrancy_rating', {})
            
            summary_data.append({
                'Repository': analysis['repository'],
                'Quality Score': round(analysis['quality_metrics']['quality_score'], 2),
                'Aberrancy Score': round(aberrancy_score, 2),
                'Industry Rating': self._get_overall_rating(analysis),
                'Risk Level': aberrancy_rating.get('risk_level', 'N/A'),
                'Risk Factors': '\n'.join(analysis['aberrancy_metrics']['risk_factors']),
                'Analyzed At': analysis['analyzed_at']
            })
        
        df = pd.DataFrame(summary_data)
        df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Format the Summary sheet
        worksheet = writer.sheets['Summary']
        worksheet.column_dimensions['A'].width = 30
        worksheet.column_dimensions['B'].width = 15
        worksheet.column_dimensions['C'].width = 15
        worksheet.column_dimensions['D'].width = 20
        worksheet.column_dimensions['E'].width = 15
        worksheet.column_dimensions['F'].width = 40
        worksheet.column_dimensions['G'].width = 25

        # Adjust row height for risk factors
        for idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            if '\n' in str(row[5].value):  # Risk Factors column
                worksheet.row_dimensions[idx].height = 15 * (str(row[5].value).count('\n') + 1)

    def _export_industry_standards(self, writer: pd.ExcelWriter):
        """Export industry standards to Excel."""
        standards_data = []
        
        # Add Aberrancy Standards
        for rating, criteria in IndustryStandards.ABERRANCY_STANDARDS.items():
            standards_data.append({
                'Category': 'Aberrancy Score',
                'Rating': rating.title(),
                'Criteria': f"Score Range: {criteria['score_range'][0]}-{criteria['score_range'][1]}",
                'Description': f"{criteria['description']} ({criteria['risk_level']})"
            })
        
        # Commit standards
        for rating, criteria in IndustryStandards.COMMIT_STANDARDS['frequency'].items():
            standards_data.append({
                'Category': 'Commit Frequency',
                'Rating': rating.title(),
                'Criteria': f"Weekly: {criteria['min_weekly']}-{criteria['max_weekly']}, Var: {criteria['variance_threshold']}",
                'Description': IndustryStandards.COMMIT_STANDARDS['description'][rating]
            })
        
        # Code churn standards
        for rating, criteria in IndustryStandards.CODE_CHURN_STANDARDS['weekly_churn'].items():
            standards_data.append({
                'Category': 'Code Churn',
                'Rating': rating.title(),
                'Criteria': f"Churn: {criteria['ratio']}, Del ratio: {criteria['deletion_ratio']}",
                'Description': IndustryStandards.CODE_CHURN_STANDARDS['description'][rating]
            })
        
        # Branch standards
        for rating, criteria in IndustryStandards.BRANCH_STANDARDS['complexity'].items():
            standards_data.append({
                'Category': 'Branch Complexity',
                'Rating': rating.title(),
                'Criteria': f"Branches: {criteria['max_branches']}, Age: {criteria['max_age_days']}d",
                'Description': IndustryStandards.BRANCH_STANDARDS['description'][rating]
            })
        
        df = pd.DataFrame(standards_data)
        df.to_excel(writer, sheet_name='Standards', index=False)
        
        # Format the Standards sheet
        worksheet = writer.sheets['Standards']
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 15
        worksheet.column_dimensions['C'].width = 30
        worksheet.column_dimensions['D'].width = 50

    def _export_metric_definitions(self, writer: pd.ExcelWriter):
        """Export metric definitions to Excel."""
        definitions_data = []
        
        # Quality Metrics
        for metric, definition in MetricDefinitions.QUALITY_METRICS.items():
            definitions_data.append({
                'Category': 'Code Quality',
                'Metric': metric,
                'Definition': definition
            })
        
        # Aberrancy Metrics
        for category, details in MetricDefinitions.ABERRANCY_METRICS.items():
            for metric, definition in details['metrics'].items():
                definitions_data.append({
                    'Category': details['title'],
                    'Metric': metric,
                    'Definition': definition
                })
        
        # Effort Metrics
        for metric, definition in MetricDefinitions.EFFORT_METRICS.items():
            definitions_data.append({
                'Category': 'Effort Analysis',
                'Metric': metric,
                'Definition': definition
            })
        
        # Ratings
        for rating, definition in MetricDefinitions.RATINGS.items():
            definitions_data.append({
                'Category': 'Ratings',
                'Metric': rating,
                'Definition': definition
            })
        
        df = pd.DataFrame(definitions_data)
        df.to_excel(writer, sheet_name='Definitions', index=False)
        
        # Format the Definitions sheet
        worksheet = writer.sheets['Definitions']
        worksheet.column_dimensions['A'].width = 25
        worksheet.column_dimensions['B'].width = 25
        worksheet.column_dimensions['C'].width = 60

    def _export_detailed_metrics(self, analyses: List[Dict], writer: pd.ExcelWriter):
        """Export detailed metrics with industry comparisons."""
        for analysis in analyses:
            repo_name = analysis['repository']
            # Create a safe sheet name (max 31 chars, no special chars)
            safe_sheet_name = (
                ''.join(c for c in repo_name if c.isalnum() or c in '_-')
                .replace('-', '_')
                [:20]  # Leave room for '_Analysis'
            )
            sheet_name = f"{safe_sheet_name}_Analysis"
            
            aberrancy = analysis['aberrancy_metrics']
            
            # Combine all metrics into one DataFrame
            metrics_data = {
                'Section': [],
                'Metric': [],
                'Value': [],
                'Industry Standard': [],
                'Rating': []
            }
            
            # Commit Frequency Metrics
            self._add_metric_section(
                metrics_data,
                'Commit Frequency',
                aberrancy['assessment_details']['commit_frequency'],
                [
                    ('Weekly Average', 'your_weekly_avg', 'industry_avg'),
                    ('Variance', 'your_variance', 'industry_variance_threshold')
                ]
            )
            
            # Code Churn Metrics
            self._add_metric_section(
                metrics_data,
                'Code Churn',
                aberrancy['assessment_details']['code_churn'],
                [
                    ('Weekly Churn', 'your_weekly_churn', 'industry_threshold'),
                    ('Deletion Ratio', 'your_deletion_ratio', 'industry_deletion_ratio')
                ]
            )
            
            # Branch Complexity Metrics
            self._add_metric_section(
                metrics_data,
                'Branch Complexity',
                aberrancy['assessment_details']['branch_patterns'],
                [
                    ('Branch Count', 'your_branch_count', 'industry_max_branches'),
                    ('Max Branch Age', 'your_max_age', 'industry_max_age')
                ]
            )
            
            # Create DataFrame and write to Excel
            df = pd.DataFrame(metrics_data)
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Format the sheet
            worksheet = writer.sheets[sheet_name]
            worksheet.column_dimensions['A'].width = 20
            worksheet.column_dimensions['B'].width = 20
            worksheet.column_dimensions['C'].width = 15
            worksheet.column_dimensions['D'].width = 20
            worksheet.column_dimensions['E'].width = 15
            
            # Add recommendations section
            start_row = len(df) + 3
            recommendations = pd.DataFrame({
                'Section': ['Recommendations'],
                'Details': [self._get_combined_recommendations(aberrancy['assessment_details'])]
            })
            recommendations.to_excel(
                writer,
                sheet_name=sheet_name,
                startrow=start_row,
                index=False
            )

    def _add_metric_section(self, metrics_data: Dict, section: str, assessment: Dict, metric_mappings: List[tuple]):
        """Add a section of metrics to the metrics data dictionary."""
        comp = assessment.get('industry_comparison', {})
        
        # Get industry standards based on section
        industry_standards = self._get_industry_standards(section)
        
        for metric_name, your_key, industry_key in metric_mappings:
            metrics_data['Section'].append(section)
            metrics_data['Metric'].append(metric_name)
            metrics_data['Value'].append(comp.get(your_key, 'N/A'))
            
            # Always show industry standard, even if repo metrics couldn't be calculated
            industry_value = comp.get(industry_key)
            if industry_value is None:
                industry_value = industry_standards.get(metric_name, 'N/A')
            metrics_data['Industry Standard'].append(industry_value)
            
            metrics_data['Rating'].append(comp.get('rating', 'Below Average').title())

    def _get_industry_standards(self, section: str) -> Dict[str, str]:
        """Get industry standards for a given section."""
        if section == 'Commit Frequency':
            return {
                'Weekly Average': '3-15 commits',
                'Variance': '< 5 (excellent), < 10 (good)'
            }
        elif section == 'Code Churn':
            return {
                'Weekly Churn': '< 200 (excellent), < 500 (good)',
                'Deletion Ratio': '< 0.8 (excellent), < 1.0 (good)'
            }
        elif section == 'Branch Complexity':
            return {
                'Branch Count': '< 5 (excellent), < 8 (good)',
                'Max Branch Age': '< 7 days (excellent), < 14 days (good)'
            }
        return {}

    def _get_combined_recommendations(self, assessment_details: Dict) -> str:
        """Combine all recommendations into a single formatted string."""
        recommendations = []
        for section, details in assessment_details.items():
            section_recs = details.get('recommendations', [])
            if section_recs:
                section_name = section.replace('_', ' ').title()
                recommendations.append(f"{section_name}:")
                recommendations.extend([f"- {rec}" for rec in section_recs])
        return '\n'.join(recommendations)

    def _format_excel_sheets(self, writer: pd.ExcelWriter):
        """Apply final formatting to all sheets."""
        workbook = writer.book
        for worksheet in workbook.worksheets:
            # Format header row
            for cell in worksheet[1]:
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(
                    start_color='CCE5FF',
                    end_color='CCE5FF',
                    fill_type='solid'
                )
            
            # Add borders
            thin_border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )
            
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    
            # Adjust row height
            for row in worksheet.rows:
                worksheet.row_dimensions[row[0].row].height = 15

    def get_code_frequency_stats(self, repo: str) -> List[List[int]]:
        """Get code frequency statistics."""
        url = f'{self.base_url}/repos/{self.org}/{repo}/stats/code_frequency'
        try:
            response = requests.get(url, headers=self.headers)
            response.raise_for_status()
            return response.json()
        except Exception as e:
            logging.error(f"Error fetching code frequency stats for {repo}: {str(e)}")
            return []

    def get_branches(self, repo: str) -> List[Dict]:
        """Get repository branches."""
        url = f'{self.base_url}/repos/{self.org}/{repo}/branches'
        try:
            response = requests.get(url, headers=self.headers)
            response.raise_for_status()
            return response.json()
        except Exception as e:
            logging.error(f"Error fetching branches for {repo}: {str(e)}")
            return []

    def get_branch_last_commit(self, repo: str, branch: str) -> Dict:
        """Get last commit information for a branch."""
        url = f'{self.base_url}/repos/{self.org}/{repo}/branches/{branch}'
        try:
            response = requests.get(url, headers=self.headers)
            response.raise_for_status()
            data = response.json()
            
            # Handle case where commit data might be missing or malformed
            if not isinstance(data, dict):
                logging.warning(f"Invalid response data structure for {repo}/{branch}")
                return {'commit': {'committer': {'date': None}}}
            
            if 'commit' not in data:
                logging.warning(f"Missing commit data for {repo}/{branch}")
                return {'commit': {'committer': {'date': None}}}
            
            commit_data = data['commit']
            if not isinstance(commit_data, dict):
                logging.warning(f"Invalid commit data type for {repo}/{branch}")
                return {'commit': {'committer': {'date': None}}}
            
            if 'committer' not in commit_data:
                logging.warning(f"Missing committer data for {repo}/{branch}")
                commit_data['committer'] = {'date': None}
            
            committer = commit_data['committer']
            if not isinstance(committer, dict):
                logging.warning(f"Invalid committer data type for {repo}/{branch}")
                commit_data['committer'] = {'date': None}
            elif 'date' not in committer:
                logging.warning(f"Missing date in committer data for {repo}/{branch}")
                committer['date'] = None
            
            return data
        except Exception as e:
            logging.error(f"Error fetching branch info for {repo}/{branch}: {str(e)}")
            return {'commit': {'committer': {'date': None}}}

    def _get_overall_rating(self, analysis: Dict) -> str:
        """Calculate overall rating based on all metrics."""
        score = analysis['quality_metrics']['quality_score']
        if score >= 90:
            return 'Excellent (Top 10%)'
        elif score >= 80:
            return 'Good (Top 25%)'
        elif score >= 70:
            return 'Average (Top 50%)'
        else:
            return 'Below Average'

def main():
    """Main function to run the analysis."""
    # Load environment variables
    load_dotenv()
    
    # Get required environment variables
    token = os.getenv('GITHUB_TOKEN')
    org = os.getenv('GITHUB_ORG')
    
    # Get repository configuration
    repos_str = os.getenv('GITHUB_REPOS')
    repos_file = os.getenv('GITHUB_REPOS_FILE')
    
    repos = []
    
    # Try to get repos from GITHUB_REPOS environment variable
    if repos_str:
        repos.extend([repo.strip() for repo in repos_str.split(',') if repo.strip()])
    
    # Try to get repos from GITHUB_REPOS_FILE
    if repos_file and os.path.exists(repos_file):
        try:
            with open(repos_file, 'r') as f:
                repos.extend([line.strip() for line in f if line.strip()])
        except Exception as e:
            logging.error(f"Error reading repos file: {str(e)}")
    
    if not all([token, org]):
        logging.error("Missing required environment variables: GITHUB_TOKEN and GITHUB_ORG")
        return
    
    if not repos:
        logging.error("No repositories specified. Set either GITHUB_REPOS or GITHUB_REPOS_FILE in .env")
        return
    
    analyzer = CodeQualityAnalyzer(token, org)
    analyses = []
    
    # Remove duplicates while preserving order
    repos = list(dict.fromkeys(repos))
    
    for repo in repos:
        logging.info(f"Analyzing repository: {repo}")
        analysis = analyzer.analyze_repository(repo)
        analyses.append(analysis)
    
    # Export results
    output_file = f"{org}_code_quality_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    analyzer.export_to_excel(analyses, output_file)
    logging.info(f"Analysis completed and exported to {output_file}")

if __name__ == "__main__":
    main() 